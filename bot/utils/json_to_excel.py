import openpyxl
import json


def get_response_as_excel(r_json):
    workbook = openpyxl.Workbook()
    workbook.save('responses.xlsx')
    workbook = openpyxl.load_workbook('responses.xlsx')
    ws = workbook['Sheet']
    for i, v in enumerate(json.loads(r_json[0][0])['questions']['questions']):
        v = tuple(v.keys())[0]
        ws.cell(row=1, column=i+2).value = v
    for i, jsn in enumerate(r_json):
        jsn = json.loads(jsn[0])
        ws.cell(row=i+2, column=1).value = jsn['user']
        for j, v in enumerate(jsn['responses']):
            v = tuple(v.values())[0]
            ws.cell(row=i+2, column=j+2).value = v
    workbook.save('responses.xlsx')
    workbook.close()
    return(open('responses.xlsx', 'rb'))